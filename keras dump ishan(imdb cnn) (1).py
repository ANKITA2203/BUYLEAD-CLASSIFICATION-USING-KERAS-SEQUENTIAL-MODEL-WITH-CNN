#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jun 21 11:43:46 2018

@author: imart
"""

import numpy
import pandas
import keras
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from keras.preprocessing.text import Tokenizer
from keras.models import Sequential
from keras.layers import Dense
from keras.wrappers.scikit_learn import KerasClassifier
from keras.utils import np_utils
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import KFold
from sklearn.preprocessing import LabelEncoder
from sklearn.pipeline import Pipeline
from sklearn.cross_validation import train_test_split
from keras.layers import Flatten
from keras.layers.convolutional import Conv1D
from keras.layers.convolutional import MaxPooling1D
from keras.layers.embeddings import Embedding
from keras.preprocessing import sequence

from keras.preprocessing import sequence
from keras.models import Sequential
from keras.layers import Dense, Dropout, Activation
from keras.layers import Embedding
from keras.layers import Conv1D, GlobalMaxPooling1D
import math


max_features = 4000
maxlen = 1000
batch_size = 256
embedding_dims = 50
filters = 250
kernel_size = 3
hidden_dims = 250
epochs = 2


keras_wb = load_workbook('keras dump ishan.xlsx')
keras_wb_sheet1 = keras_wb['Sheet1']
keras_rows = keras_wb_sheet1.max_row
print(keras_rows)

x1=[]
X=[]
y=[]
y1=[]
for i in range(1,keras_rows+1):
    bl = keras_wb_sheet1.cell(column=1,row=i).value.lower()
    sold = keras_wb_sheet1.cell(column=2,row=i).value.lower()
    sold1 = int(keras_wb_sheet1.cell(column=2,row=i).value.lower())
    x1.append(bl)
    y.append(sold)
    y1.append(sold1)
    
    
    
seed = 7
numpy.random.seed(seed)
t = Tokenizer(1000)

t.fit_on_texts(x1)

#print(t.word_counts)
#print(t.document_count)
#print(t.word_index)
#print(t.word_docs)

X = t.texts_to_matrix(x1, mode='tfidf')
encoder = LabelEncoder()
encoder.fit(y)
encoded_Y = encoder.transform(y)
dummy_y = np_utils.to_categorical(encoded_Y)

model = Sequential()

# we start off with an efficient embedding layer which maps
# our vocab indices into embedding_dims dimensions
model.add(Embedding(max_features,
                    embedding_dims,
                    input_length=maxlen))
model.add(Dropout(0.2))

# we add a Convolution1D, which will learn filters
# word group filters of size filter_length:
model.add(Conv1D(filters,
                 kernel_size,
                 padding='valid',
                 activation='relu',
                 strides=1))
# we use max pooling:
model.add(GlobalMaxPooling1D())

# We add a vanilla hidden layer:
model.add(Dense(hidden_dims))
model.add(Dropout(0.2))
model.add(Activation('relu'))

# We project onto a single unit output layer, and squash it with a sigmoid:
model.add(Dense(2))
model.add(Activation('softmax'))

model.compile(loss='binary_crossentropy',
              optimizer='adam',
              metrics=['accuracy'])

kfold = KFold(n_splits=10, shuffle=True, random_state=seed)

X_train, X_test, Y_train, Y_test = train_test_split(X, dummy_y, test_size=0.3, random_state=seed)

model.fit(X_train, Y_train,
          batch_size=batch_size,
          epochs=epochs,
          validation_data=(X_test, Y_test))


'''

predictions = model.predict(X)
print(predictions)

probabilities = model.predict_proba(X)
print(probabilities)


from sklearn.metrics import confusion_matrix

from sklearn.metrics import accuracy_score

from sklearn.metrics import classification_report

actual = y1
#print(Y_test)
print (actual)
print (len(actual))

predictions1 = predictions.astype(int)
predicted = (predictions1)

print(predicted)
print(len(predicted))


results = confusion_matrix(actual, predicted)

print ('Confusion Matrix :')

print(results)

#print ('Accuracy Score :'),accuracy_score(actual,predicted)

print ('Report : ')

print (classification_report(actual, predicted))
'''
