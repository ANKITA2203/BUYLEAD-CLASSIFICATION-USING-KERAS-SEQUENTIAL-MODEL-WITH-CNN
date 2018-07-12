# -*- coding: utf-8 -*-
"""
Created on Wed Jun 13 15:47:47 2018

@author: hemlata
"""

import numpy
import pandas
import keras
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from keras.preprocessing.text import Tokenizer
from keras.models import Sequential
from keras.layers import Dense
from keras.wrappers.scikit_learn import KerasClassifier
from keras.utils import np_utils
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import KFold
from sklearn.preprocessing import LabelEncoder
from sklearn.pipeline import Pipeline
from sklearn.cross_validation import train_test_split
from keras.layers import Flatten
from keras.layers.convolutional import Conv1D
from keras.layers.convolutional import MaxPooling1D
from keras.layers.embeddings import Embedding
from keras.preprocessing import sequence

keras_wb = load_workbook('Keras dump ishan.xlsx')
keras_wb_sheet1 = keras_wb['Sheet1']
keras_rows = keras_wb_sheet1.max_row
print(keras_rows)

x1=[]
X=[]
y=[]
y1=[]
for i in range(1,keras_rows+1):
    bl = keras_wb_sheet1.cell(column=1,row=i).value.lower()
    sold = keras_wb_sheet1.cell(column=2,row=i).value.lower()
    sold1 = int(keras_wb_sheet1.cell(column=2,row=i).value.lower())
    x1.append(bl)
    y.append(sold)
    y1.append(sold1)

#print(x1)
#print(y)

'''
for j in range(0,keras_rows):
    text = x1[j]
    # estimate the size of the vocabulary
    words = set(text_to_word_sequence(text))
    vocab_size = len(words)
    print(vocab_size)
    # integer encode the document
    result = one_hot(text, round(vocab_size*1.3))
    x.append(result)
    print(result)
'''

seed = 7
numpy.random.seed(seed)
t = Tokenizer(num_words=1000)

t.fit_on_texts(x1)

print(t.word_counts)
print(t.document_count)
print(t.word_index)
print(t.word_docs)

X = t.texts_to_matrix(x1, mode='tfidf')
encoder = LabelEncoder()
encoder.fit(y)
encoded_Y = encoder.transform(y)
dummy_y = np_utils.to_categorical(encoded_Y)

def baseline_model():
	# create model
    model = Sequential()
    model.add(Embedding(500, 1000, input_length=1000))
    model.add(Conv1D(100,kernel_size=3, padding='same', activation='relu'))
    model.add(MaxPooling1D(pool_size=2))
    model.add(Flatten())
    model.add(Dense(2, activation='softmax'))
    model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
    keras.optimizers.Adam(lr=0.001, beta_1=0.9, beta_2=0.999, epsilon=None, decay=0.0, amsgrad=False)
    return model


estimator = KerasClassifier(build_fn=baseline_model, nb_epoch=10, batch_size=5, verbose=0)

kfold = KFold(n_splits=10, shuffle=True, random_state=seed)

X_train, X_test, Y_train, Y_test = train_test_split(X, dummy_y, test_size=0.33, random_state=seed)
'''print (len(dummy_y))
print(len(X_train))'''

estimator.fit(X,dummy_y)

predictions = estimator.predict(X)
probabilities = estimator.predict_proba(X)

print(predictions)
print(probabilities)

#print(encoder.inverse_transform(predictions))


from sklearn.metrics import confusion_matrix

from sklearn.metrics import accuracy_score

from sklearn.metrics import classification_report

actual = y1
#print(Y_test)
print (actual)
print (len(actual))

predicted = (predictions)
print(predicted)
print(len(predicted))


results = confusion_matrix(actual, predicted)

print ('Confusion Matrix :')

print(results)

#print ('Accuracy Score :'),accuracy_score(actual,predicted)

print ('Report : ')

print (classification_report(actual, predicted))

'''
keras_wb = load_workbook('kddcnn.xlsx')
keras_wb_sheet1 = keras_wb['Sheet1']
keras_wb_sheet1 = keras_wb.active

for i in range(1,keras_rows+1):
    keras_wb_sheet1.cell(column=1,row=i).value = x1[i-1]
    keras_wb_sheet1.cell(column=2,row=i).value = y1[i-1]
    keras_wb_sheet1.cell(column=3,row=i).value = predictions[i-1]
    keras_wb_sheet1.cell(column=4,row=i).value = probabilities[i-1][0]
    keras_wb_sheet1.cell(column=5,row=i).value = probabilities[i-1][1]
    
keras_wb.save('kddcnn.xlsx')'''